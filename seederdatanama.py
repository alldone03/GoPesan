# import pandas as pd

# # Load the xlsx file
# excel_data = pd.read_excel('DATANRPNAMA.xlsx')
# # Read the values of the file in the dataframe
# data = pd.DataFrame(excel_data, columns=[
#                     'No', 'NRP', 'NAMA'])
# # Print the content
# print("The content of the file is:\n", data)


# Import openyxl module
import openpyxl
import requests


# Define variable to load the wookbook
wookbook = openpyxl.load_workbook("DATANRPNAMA.xlsx")

# Define variable to read the active sheet:
worksheet = wookbook.active

# Iterate the loop to read the cell values
for i in range(0, worksheet.max_row):
    for col in range(0, worksheet.max_column):
        # print(i+65)
        # print(chr(col+65)+str(i))
        # print(worksheet[chr(65)+str(i+1)].value)
        datanrp = worksheet[chr(65)+str(i+1)].value
        datanama = worksheet[chr(66)+str(i+1)].value
    # for col in worksheet.iter_cols(1, worksheet.max_column):
    # print(col[i].value, end="\t\t")
    # print(str(col)+'\n')
    print(str(datanrp) + " " + str(datanama))
    x = requests.get(
        'http://192.168.1.102/api/updatename?username='+str(datanrp)+'&nama='+str(datanama))
    print(x.content)
    print('')
